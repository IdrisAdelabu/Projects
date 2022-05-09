from openpyxl import load_workbook

workbook = load_workbook(filename = "C:\\Users\\gbade\\Documents\\shortest_path.xlsx")
workbook.active = 0
sheet = workbook.active

num_nodes = sheet.max_column

for trial_node in range(1, num_nodes + 1):
    for y in range(1, num_nodes + 1):
        rand_list = [i.value for i in sheet["y"]]
        rand_list_copy = [i.value for i in sheet["y"]]
        emp_list = []
        for path in rand_list_copy:

            value_count = len(emp_list)
            num = rand_list.index(path) + value_count
            emp_list.append(path)
            rand = sheet.cell(row=y, column=num + 1).value
            tentative_value = sheet.cell(row=y, column=trial_node).value + sheet.cell(row=trial_node,
                                                                                      column=num + 1).value
            if tentative_value < rand:
                sheet.cell(row=y, column=num + 1).value = tentative_value
                workbook.active = 1
                sheet = workbook.active
                sheet.cell(row=y, column=num + 1).value = trial_node

            workbook.active = 0
            sheet = workbook.active
            rand_list.pop(rand_list.index(path))

workbook.save(filename="shortest_path.xlsx")