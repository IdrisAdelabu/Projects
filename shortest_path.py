import os
from openpyxl import load_workbook

workbook = load_workbook(filename = "C:\\Users\\gbade\\Documents\\shortest_path.xlsx")
workbook.active = 0
sheet = workbook.active
perm_dict = {}
temp_dict = {y:[x.value,1] for y,x in zip(range(1, sheet.max_column+1), [x for x in sheet["1"]])}
temp_dict.pop(1)

while temp_dict != {}:

    path_list = [pa for [pa,node] in list(temp_dict.values())]
    key_list = list(temp_dict.keys())
    min_path = min(path_list)
    min_path_index = path_list.index(min_path)
    selected_node = key_list[min_path_index]
    add_node = temp_dict[selected_node]
    perm_dict[selected_node] = add_node
    temp_dict.pop(selected_node)

    for x in key_list:
        if x != selected_node:
            current_value = temp_dict[x][0]
            new_value = perm_dict[selected_node][0] + sheet.cell(row=selected_node, column=x).value
            if new_value < current_value:
                temp_dict[x][0] = new_value
                temp_dict[x][1] = selected_node

for desired_node in range(2,sheet.max_column+1):

    shortest_path = perm_dict[desired_node][0]
    path_route = [desired_node]
    starting_node = perm_dict[desired_node][1]

    while starting_node != 1:
        path_route.append(starting_node)
        starting_node = perm_dict[starting_node][1]

    path_route.append(1)
    path_route.reverse()
    new_str = [str(n) for n in path_route]
    path_str = "-->".join(new_str)
    print(f"The shortest path to node {desired_node} is: {shortest_path}")
    print(f"The path route is {path_str}")